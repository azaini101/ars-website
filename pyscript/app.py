from flask import Flask, send_file, request
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import date
import os
import io
from dotenv import load_dotenv
load_dotenv()
link = os.environ.get("MONGODB_URI")
app = Flask(__name__)

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected
        

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def boldHeader(sh, col):
    for i in range(1, col+1):
        cell = sh.cell(row = 1, column = i)
        cell.font = Font(bold=True)

def iterateCursor(fields, cursor, sh, wb, sheets):
    for document in cursor:
        row = sh.max_row + 1
        col = 1
        pasteRange(1, 1, len(fields), 1, sh, [fields])
        boldHeader(sh, len(fields))
        for field in fields:
            cell = sh.cell(row = row, column = col)
            try:
                val = document[field]
            except KeyError:
                continue
            if(field == "phone" or field == "emergencyPhone"):
                try:
                    val = float(val)
                except ValueError:
                    pass
            cell.value = val
            col += 1
        copied = copyRange(1, row, 11, row, sh)

        for key in sheets:
            if(document["services"].find(key) != -1):
                try:
                    sersh = wb[key]
                except:
                    sersh = wb.create_sheet(key)
                    pasteRange(1, 1, len(fields), 1, sersh, [fields])
                    boldHeader(sersh, len(fields))
                row = sersh.max_row + 1
                pasteRange(1, row, 11, row, sersh, copied)
    return fields


@app.route("/getData", methods=['GET'])
def getData():
    dataType = request.args.get('dataType')
    print(dataType)
    return main(dataType)

def main(dataType):
    today = date.today()
    the_date = today.strftime("%b_%d_%Y")

    client = MongoClient(link)
    print("got it")
    db = client['donationsDB']
    collection = db[dataType]
    cursor = collection.find({})

    wb = Workbook()
    sheets = {
    "Housing",
    "Employment",
    "English",
    "Benefits",
    "Mentoring",
    "Medical",
    "Legal",
    "Education",
    "Transportation",
    "Financial",
    "Electronics",
    "Apartment Setup"
    }

    sh = wb.active
    sh.title = "Signups_{}".format(the_date)

    fields = [
        "firstName",
        "lastName",
        "email",
        "phone",
        "emergencyFirstName",
        "emergencyLastName",
        "emergencyPhone",
        "services",
        "times",
        "languages",
        "notes",
        "idaraMember",
        "idaraVisits",
        "faith"    ]
    iterateCursor(fields, cursor, sh, wb, sheets)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    if(dataType == "donations"):
        dataType = "Idara"
    file_name = "{}_{}_export.xlsx".format(the_date, dataType)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=file_name, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)